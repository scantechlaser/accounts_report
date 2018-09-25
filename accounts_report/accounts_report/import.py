import frappe
from frappe.model.document import Document
from frappe import _
from frappe.utils.response import build_response
import frappe.utils.user
import requests
import time as t
from frappe.utils import getdate, validate_email_add, today
from frappe.utils import cstr, flt, cint, getdate, now_datetime, formatdate, strip
import traceback
from datetime import *
import sys
import openpyxl
import re
from openpyxl import Workbook


@frappe.whitelist(allow_guest=True)
def importDeliveryNoteItem():
	r={}
	F=[]
	ImportData=""
	duplicatedata={}
	RateDict={}
	RateList=[]
	RateLambda={}
	RateData={}
	diagnose = []

	insurance_id=""
	SrNo=""
	ServiceName=""
	ServiceCode=""
	ServiceCharge=""

	count=0
	temp_rcount=0
	service_id=""



	f=[]
	ErrorCell={}
	ServiceErrorCell={}
	msgErrorCell=""
	cellvalues=[]
	countValue=0
	d=""
	category = ''
	unit = ''
	vendor = ''
	brandNameQuery = ''
	sfda = ''
	quanrity_per_sale_unit = ''
	price = ''
	date=now_datetime()
	finalCount = 0
	# tpaCode	tpaName	contactPerson	contactAddress	contactNumber

	Rate=['Name','ItemCode', 'Qty', 'Description']
	try:
		if True:
			# file_path = '/home/frappe/frappe-bench/sites/assets/Templates/CompleteServicesnew1.xlsx'
			file_path = '/home/frappe/frappe-bench/sites/assets/accounts_report/delivery_challan.xlsx'
			# frappe.throw(_(file_path))

			if file_path:
				diagnose.append(file_path)
				wb=openpyxl.load_workbook(file_path)


				sheet_name=wb.get_sheet_names()
				# frappe.throw(_(sheet_name))
				# sheet=wb.get_sheet_by_name("Human-Drugs")
				# sheet=wb.get_sheet_by_name("Herbal-Drugs")
				sheet=wb.get_sheet_by_name("Delivery Challan")


				row_count=sheet.max_row


				column_count=sheet.max_column
				

				for i in range(1,row_count+1):
					# diagnose.append(i)
					cell_0=sheet.cell(row=i,column=1).value
					cell_1=sheet.cell(row=i,column=2).value
					cell_2=sheet.cell(row=i,column=3).value
					
					if cell_0 and cell_1 :

						temp_rcount=temp_rcount+1
						
						f.append(i)
						
					else:

						frappe.local.response['data']=[]
						frappe.local.response['msg']="Please provide mandatory fields"
						frappe.local.response['status']=False
						return False

				

				for x in range(1,2):
					diagnose.append(count)
					if count>0:
						break
					else:
						for y in range(1,4):
							cell_data=(sheet.cell(row=x,column=y).value)
							#cellvalues.append({"value":cell_data})
							if cell_data!=Rate[y-1]:
								count=count+1
								msgErrorCell="Column names not matching"
								break

				
				if count==0:
					# login_website(webUser,webPassword)
					diagnose.append({"count" : count})
					for i in range(2,temp_rcount+1):
						tpaCode	= ''
						tpaName	= ''
						contactPerson =	''
						contactAddress = ''
						contactNumber = ''
						Description = ''
						# frappe.msgprint("i")
						# frappe.msgprint(i)
						if countValue>0:
							break
						else:
							for j in range(1,5):
								
								cell_data=sheet.cell(row=i,column=j).value
								d=d+str(j)
								if cell_data=="None":
									countValue=countValue+1
									ErrorCell={"Row":f[i-1],"Column":j}
									break
								else:

									if j==1:
										cell_data=sheet.cell(row=i,column=j).value
										
										gettpaCode=cell_data
										if gettpaCode:
											tpaCode=sheet.cell(row=i,column=j).value

										
									elif j==2:
										cell_data=sheet.cell(row=i,column=j).value
										gettpaName=cell_data
										if gettpaName:
											tpaName=sheet.cell(row=i,column=j).value
											
									elif j==3:
										cell_data = sheet.cell(row=i,column=j).value
										getcontactPerson = cell_data
										if getcontactPerson:
											contactPerson=sheet.cell(row=i,column=j).value
									elif j==4:
										cell_data = sheet.cell(row=i,column=j).value
										getDescription = cell_data
										if getDescription:
											Description=sheet.cell(row=i,column=j).value


							# tpaCode	tpaName	contactPerson	contactAddress	contactNumber
							RateList.append({
								"challanName":tpaCode,
								"itemCode":tpaName,
								"qty":contactPerson,
								"Description":Description
								})


					for s in RateList:

						# frappe.msgprint(s['itemCode'])

						uom = ''
						base_rate = 0.0
						hsn = ''

						my_sql = "SELECT stock_uom, gst_hsn_code FROM `tabItem` WHERE item_code	= '"+str(s['itemCode'])+"'"
						getUOM = frappe.db.sql(my_sql, as_dict=True)
						if getUOM:

							uom = getUOM[0].stock_uom
							hsn = getUOM[0].gst_hsn_code

						my_sql = "SELECT I.base_rate AS base_rate FROM `tabPurchase Receipt Item` I RIGHT JOIN `tabPurchase Receipt` R ON R.name = I.parent WHERE I.`item_code` = '"+str(s['itemCode'])+"' ORDER BY R.posting_date DESC LIMIT 0,1"
						getBaseRate =frappe.db.sql(my_sql, as_dict=True)
						if getBaseRate:

							base_rate = getBaseRate[0].base_rate

						excelQty = [float(k) for k in re.findall(r'-?\d+\.?\d*', str(s['qty']))]

						amount = float(excelQty[0]) * float(base_rate)

						my_sql = "SELECT * FROM `tabDelivery Challan Item` WHERE parent = '"+str(s['challanName'])+"' and item_code = '"+str(s['itemCode'])+"'"
						getChallanItem = frappe.db.sql(my_sql, as_dict=True)
						if getChallanItem:

							for o in getChallanItem:

								itemQty = [float(k) for k in re.findall(r'-?\d+\.?\d*', str(o['qty']))]


								if(str(itemQty[0]) == str(excelQty[0])):

									my_sql = "UPDATE `tabDelivery Challan Item` SET qty = '"+str(itemQty[0])+"', unit = '"+str(uom)+"', amount = '"+str(amount)+"', rate = '"+str(base_rate)+"' WHERE name ='"+str(o.name)+"'"
									frappe.db.sql(my_sql)
									finalCount = finalCount + 1
						else:
							frappe.msgprint(str(s['itemCode'])+'----'+str(s['challanName']))
					
					
					if ErrorCell:
						# frappe.throw(_(sheet.cell(row=int(ErrorCell['Row'])+1,column=int(ErrorCell['Column'])).value))
						msgErrorCell="Invalid Template,Please check Column='"+str(ErrorCell['Column'])+"' and Row='"+str(ErrorCell['Row'])+"'"
					if ServiceErrorCell:
						msgErrorCell="Service not available with '"+str(ServiceErrorCell['serviceName'])+"' name,Please check Column='"+str(ServiceErrorCell['Column'])+"' and Row='"+str(ServiceErrorCell['Row'])+"'"
					if countValue>0:
						frappe.local.response['data']=ErrorCell
						frappe.local.response['message']=msgErrorCell
						frappe.local.response['status']=False
					else:
						frappe.local.response['data']=finalCount
						frappe.local.response['len']=len(RateList)
						frappe.local.response['message']="Data Uploading In Progress.... Check after 15 minutes"
						frappe.local.response['message']="Data Uploading In Progress.... Check after a while"
						frappe.local.response['message']="Data Uploaded Successfully"
						frappe.local.response['status']=True
				else:
					frappe.local.response['data']=ErrorCell
					frappe.local.response['message']=msgErrorCell#"Invalid template"
					frappe.local.response['status']=False

		else:
			frappe.local.response['data']=''
			frappe.local.response['message']="Debug not active"
			frappe.local.response['status']=False
		frappe.local.response['diagnose']=diagnose

	except:
		# frappe.msgprint(sql)
		# frappe.msgprint(i)
		type_, value_, traceback_ = sys.exc_info()
		debug=1
		if debug == 1:
			frappe.local.response['data'] = {"type":str(type_),"value":str(value_),"stack":str(traceback.format_tb(traceback_))}
		else:
			frappe.local.response['data'] = ""
		frappe.local.response['message'] = "Error!!!"
		frappe.local.response['status'] = False


